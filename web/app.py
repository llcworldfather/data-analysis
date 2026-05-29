"""
BOM 成本分析 Web 应用
======================
自包含：与本目录下 process_excel.py、templates/、requirements.txt 即可运行。

启动（请在 web 目录下，或 python 路径指向本目录下的 app.py）:
  双击 start_web.bat
  或: python app.py

访问: http://127.0.0.1:5000（约 1 秒后尝试自动打开浏览器）
不自动打开: 环境变量 BOM_WEB_NO_BROWSER=1
"""
from __future__ import annotations

import io
import json
import math
import os
import re
import threading
import time
import uuid
import traceback
import webbrowser
import zipfile
import html
import urllib.error
import urllib.parse
import urllib.request

import requests
from pathlib import Path
from datetime import datetime
import sys

from flask import Flask, request, jsonify, send_file, render_template, Response, stream_with_context
import pandas as pd
import polars as pl

from services import bi_cache
from routes.bi import bp as bi_bp, init_blueprint as init_bi_blueprint
from routes.bi_common import optional_rounded_float

from process_excel import (
    prepare,
    aggregate_lines,
    enrich,
    sheet_product_summary,
    sheet_bom_detail,
    sheet_component_global_rank,
    sheet_cross_product_volatility,
    top3_per_product,
    sheet_price_history,
    normalize_columns,
    norm_material_code,
    read_excel_as_polars,
    read_price_excel,
    REQUIRED_COLS,
    COL_PRODUCT,
    COL_COMPONENT,
    COL_QTY,
    COL_UNIT,
    COL_UNIT_PRICE,
    COL_MAKTX,
    COL_CREATEDATE,
    COL_CATEGORY,
    _write_sheet,
)

app = Flask(__name__, template_folder="templates", static_folder="static", static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

WEB_DIR    = Path(__file__).parent
UPLOAD_DIR = WEB_DIR / "uploads"
RESULT_DIR = WEB_DIR / "results"
UPLOAD_DIR.mkdir(exist_ok=True)
RESULT_DIR.mkdir(exist_ok=True)
bi_cache.init(RESULT_DIR)


def _retention_days() -> int:
    try:
        return max(1, int(os.environ.get("BOM_RETENTION_DAYS", "7")))
    except ValueError:
        return 7


def purge_stale_files(directory: Path, *, max_age_days: int | None = None) -> int:
    """删除 directory 下 mtime 早于保留期的普通文件。"""
    if not directory.is_dir():
        return 0
    days = _retention_days() if max_age_days is None else max(1, max_age_days)
    cutoff = time.time() - days * 86400
    removed = 0
    for p in directory.iterdir():
        if p.is_file() and p.stat().st_mtime < cutoff:
            p.unlink(missing_ok=True)
            removed += 1
    return removed


def _purge_upload_and_results() -> None:
    u = purge_stale_files(UPLOAD_DIR)
    r = purge_stale_files(RESULT_DIR)
    if u or r:
        app.logger.info(
            "已清理过期文件 uploads=%s results=%s（保留 %s 天）",
            u,
            r,
            _retention_days(),
        )


def _schedule_daily_purge() -> None:
    def _tick() -> None:
        _purge_upload_and_results()
        threading.Timer(86400.0, _tick).start()

    threading.Timer(86400.0, _tick).start()


def _wait_result_excel(session_id: str, file_path: Path, timeout: float = 120.0) -> bool:
    """等待后台 Excel 写完；已存在则立即返回 True。"""
    if file_path.exists():
        return True
    with _EXCEL_EVENTS_LOCK:
        ev = _EXCEL_EVENTS.get(session_id)
    if ev is not None:
        ev.wait(timeout=timeout)
    return file_path.exists()


def _load_env_file(path: Path) -> None:
    """加载本地 .env，已存在的系统环境变量优先级更高。"""
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


_load_env_file(WEB_DIR / ".env")

_ROOT_PROJ = Path(__file__).resolve().parent.parent
if str(_ROOT_PROJ) not in sys.path:
    sys.path.insert(0, str(_ROOT_PROJ))
from bom_date_key import calendar_date_key_yyyymmdd as _calendar_date_key_yyyymmdd

from cost_sim_predict import (
    baseline_prices_from_price_snap,
    build_product_category_index,
    build_product_cost_history,
    build_product_price_timeline,
    latest_bom_snapshot_from_rows,
    map_legacy_predict_en_to_zh,
    map_regression_analysis_en_to_zh,
    map_sensitivity_grid_en_to_zh,
    map_sensitivity_item_en_to_zh,
    normalize_predict_dataframes,
    normalize_price_list_cost_fields,
    unit_hint_from_row,
    predict_product_price,
    _predict_product_price_legacy,
    _safe_float,
    quote_month_key_from_series,
)
from material_search import web_search_material_price

_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
)


# ---------------------------------------------------------------------------
# Global JSON error handlers – 确保 API 永远返回 JSON，不返回 HTML 调试页
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def err_404(e):
    return jsonify({"error": "接口不存在", "detail": str(e)}), 404

@app.errorhandler(405)
def err_405(e):
    return jsonify({"error": "请求方法不允许", "detail": str(e)}), 405

@app.errorhandler(413)
def err_413(e):
    return jsonify({"error": "文件过大：最大支持 100 MB"}), 413

@app.errorhandler(Exception)
def err_any(e):
    app.logger.exception("未捕获异常")
    resp: dict = {"error": "服务器内部错误，请稍后重试"}
    if app.debug:
        tb = traceback.format_exc().strip().split("\n")
        resp["detail"] = tb[-1] if tb else None
    return jsonify(resp), 500


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# API – Template download
# ---------------------------------------------------------------------------

@app.route("/api/template")
def download_template():
    """生成并返回标准 BOM Excel 模板（列格式与 SAP 导出一致）。"""
    df = pd.DataFrame({
        "ZBJNO":      ["2601070001", "2601070001", "2601070001", "2601070002", "2601070002"],
        "ZSNO":       ["2516001",    "2516001",    "2516001",    "2516002",    "2516002"],
        "IDNRK":      ["1010010004", "1010010293", "1010020025", "1010010004", "1010030011"],
        "MEINS":      ["KG",         "KG",         "KG",         "KG",         "PC"],
        "MENGE":      [600.0,        400.0,        20.0,         500.0,        30.0],
        "JGLX":       [3,            3,            1,            3,            1],
        "ZDJ":        [7.00,         5.55,         2.45,         7.00,         12.80],
        "ZMATNR":     ["8040120649", "8040120649", "8040120649", "8040120650", "8040120650"],
        "MAKTX":      ["原材料A-PE管材", "外购PE粒子", "色母粒", "原材料A-PE管材", "配件B"],
        "CREATEDATE": ["20260107",   "20260107",   "20260107",   "20260107",   "20260107"],
        "分类":        ["管材类",     "管材类",     "管材类",     "管材类",     "配件类"],
    })
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="BOM数据")
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="BOM分析模板.xlsx",
    )


# ---------------------------------------------------------------------------
# API – Upload & Preview
# ---------------------------------------------------------------------------

def _xlsx_row_count(path: Path) -> int:
    """从 xlsx 快速获取数据行数（不读全部单元格）。
    策略1：读工作表 XML 前 4096 字节，解析 <dimension ref="A1:Z15000"> 元素（< 2ms）。
    策略2：若未找到则用 openpyxl read_only 的 max_row 属性（仍比全量读快很多）。
    """
    # ── 策略1：从 XML 元数据读取 ──────────────────────────────────────────
    try:
        with zipfile.ZipFile(str(path)) as zf:
            sheet = next(
                (n for n in zf.namelist()
                 if re.match(r"xl/worksheets/sheet\d+\.xml$", n)),
                None,
            )
            if sheet:
                with zf.open(sheet) as ws:
                    head = ws.read(4096).decode("utf-8", errors="ignore")
                # 匹配 ref="A1:BZ15000"，只需结束行号
                m = re.search(r'ref="[A-Za-z$]+\d+:[A-Za-z$]+(\d+)"', head)
                if m:
                    return max(0, int(m.group(1)) - 1)   # 减去表头行
    except Exception:
        pass

    # ── 策略2：openpyxl read_only（读 XML 元数据，不加载单元格）──────────
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        nrows = (ws.max_row or 1) - 1
        wb.close()
        return max(0, nrows)
    except Exception:
        pass

    return 0


def _is_excel_filename(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith(".xlsx") or (n.endswith(".xls") and not n.endswith(".xlsx"))


def _upload_save_ext(filename: str) -> str:
    """根据上传文件名决定落盘扩展名（支持模板 .xls / .xlsx）。"""
    n = (filename or "").lower()
    if n.endswith(".xlsx"):
        return ".xlsx"
    if n.endswith(".xls"):
        return ".xls"
    return ""


def _upload_bom_path(session_id: str, idx: int) -> Path | None:
    for ext in (".xlsx", ".xls"):
        p = UPLOAD_DIR / f"{session_id}_{idx}{ext}"
        if p.exists():
            return p
    return None


def _upload_price_path(session_id: str) -> Path | None:
    for ext in (".xlsx", ".xls"):
        p = UPLOAD_DIR / f"{session_id}_price{ext}"
        if p.exists():
            return p
    return None


def _read_xlsx_as_polars(path: Path) -> pl.DataFrame:
    """读取 Excel 第一张工作表（.xlsx / .xls）；物料编码列强制 Utf8。"""
    return read_excel_as_polars(path)


def _norm_product_code_series(s: pd.Series) -> pd.Series:
    """统一成品物料号为纯数字字符串，避免 Excel 浮点列变成 '8010040004.0' 导致与 BOM 对不齐。"""
    return norm_material_code(s)


# 产品价格历史清单模板：除「总成本→产品价格」外，下列列会一并并入 BOM 行（按产品/日期对齐）
_PRICE_EXTRA_COLS = ("材料成本", "工费", "重量", "报价月份", "标杆工厂")


def _attach_product_prices_pl(raw_pl: pl.DataFrame, price_path: Path) -> pl.DataFrame:
    """将「产品价格历史清单」左并入 BOM 明细行。

    业务语义（每一行）：某 **产品** 在某一 **报价月份**、某一 **重量(kg)** 下，
    **材料成本 / 工费 / 总成本均为 元/公斤**（写入列「产品价格」= 总成本每公斤）。

    对齐优先级：**产品+报价月份+重量** → **产品+报价月份** → **产品+日期** → **仅产品**（末条补全）。
    """
    raw_norm = normalize_columns(raw_pl)
    bom = raw_norm.to_pandas()
    drop_cols = ["产品价格", *_PRICE_EXTRA_COLS]
    bom = bom.drop(columns=[c for c in drop_cols if c in bom.columns], errors="ignore")

    price = read_price_excel(price_path)
    if COL_PRODUCT not in price.columns:
        if "ZMATNR" in price.columns:
            price = price.rename(columns={"ZMATNR": COL_PRODUCT})
        elif "MATNR" in price.columns:
            price = price.rename(columns={"MATNR": COL_PRODUCT})
    if COL_PRODUCT not in price.columns:
        raise ValueError("产品价格表缺少列：产品编码（或 ZMATNR / MATNR）")

    val_col = None
    # 「总成本」优先：与产品价格历史清单模板一致
    for c in (
        "总成本",
        "产品价格",
        "标价",
        "出厂价",
        "销售价",
        "DMBTR",
        "NETPR",
        "KBETR",
        "NETWR",
    ):
        if c in price.columns:
            val_col = c
            break
    if val_col is None:
        raise ValueError(
            "产品价格表缺少金额列：总成本（或 产品价格 / 标价 / DMBTR 等）"
        )
    price = price.rename(columns={val_col: "产品价格"})

    merge_value_cols = ["产品价格"] + [c for c in _PRICE_EXTRA_COLS if c in price.columns]
    date_candidates = ("生价日期", "CREATEDATE", "生产日期", "创建日期", "日期")
    price_date_col = next((c for c in date_candidates if c in price.columns), None)

    bom[COL_PRODUCT] = _norm_product_code_series(bom[COL_PRODUCT])
    price[COL_PRODUCT] = _norm_product_code_series(price[COL_PRODUCT])
    for c in merge_value_cols:
        if c in price.columns and c not in ("报价月份", "标杆工厂"):
            price[c] = pd.to_numeric(price[c], errors="coerce")

    if "报价月份" in price.columns:
        price["报价月份"] = quote_month_key_from_series(price["报价月份"])
        price["_qm"] = price["报价月份"]
    elif price_date_col:
        price["_qm"] = quote_month_key_from_series(
            price[price_date_col].map(_calendar_date_key_yyyymmdd)
        )
    if "重量" in price.columns:
        price["重量"] = pd.to_numeric(price["重量"], errors="coerce")

    if "报价月份" in bom.columns:
        bom["报价月份"] = quote_month_key_from_series(bom["报价月份"])
        bom["_qm"] = bom["报价月份"]
    bom_date_col = COL_CREATEDATE if COL_CREATEDATE in bom.columns else None
    if "_qm" not in bom.columns and bom_date_col:
        bom["_qm"] = quote_month_key_from_series(
            bom[bom_date_col].map(_calendar_date_key_yyyymmdd)
        )
    if bom_date_col:
        bom["_dk"] = bom[bom_date_col].map(_calendar_date_key_yyyymmdd)
    if price_date_col:
        price["_dk"] = price[price_date_col].map(_calendar_date_key_yyyymmdd)
    if "重量" in bom.columns:
        bom["重量"] = pd.to_numeric(bom["重量"], errors="coerce")

    def _deduped_price(keys: list[str]) -> pd.DataFrame:
        keys = [k for k in keys if k in price.columns]
        if COL_PRODUCT not in keys:
            keys = [COL_PRODUCT, *keys]
        sort_keys = keys if keys else [COL_PRODUCT]
        out = price.sort_values(sort_keys, na_position="last").drop_duplicates(
            keys, keep="last"
        )
        cols = keys + [c for c in merge_value_cols if c in out.columns]
        return out[cols].dropna(subset=["产品价格"], how="all")

    def _coalesce_merge(base: pd.DataFrame, on_keys: list[str]) -> pd.DataFrame:
        on_keys = [k for k in on_keys if k in base.columns and k in price.columns]
        if COL_PRODUCT not in on_keys:
            return base
        use = _deduped_price(on_keys)
        if use.empty:
            return base
        m = base.merge(use, on=on_keys, how="left", suffixes=("", "_px"))
        for c in merge_value_cols:
            px = f"{c}_px"
            if c not in m.columns and px in m.columns:
                m[c] = m[px]
            elif c in m.columns and px in m.columns:
                m[c] = m[c].fillna(m[px])
            m = m.drop(columns=[px], errors="ignore")
        return m

    merged = bom.copy()
    merge_chain: list[list[str]] = []
    # 修复（问题15）：对齐优先级改为「精确日期(YYYYMMDD) → 报价月份(YYYYMM) → 仅产品」。
    # 原逻辑月度(_qm)先于日期(_dk)，导致粗粒度月度匹配先填充后，
    # 精确日期匹配被 coalesce 跳过，丢失更精细的对齐机会。
    if "_dk" in merged.columns and "_dk" in price.columns:
        if "重量" in merged.columns and "重量" in price.columns:
            merge_chain.append([COL_PRODUCT, "_dk", "重量"])
        merge_chain.append([COL_PRODUCT, "_dk"])
    if "_qm" in merged.columns and "_qm" in price.columns:
        if "重量" in merged.columns and "重量" in price.columns:
            merge_chain.append([COL_PRODUCT, "_qm", "重量"])
        merge_chain.append([COL_PRODUCT, "_qm"])
    merge_chain.append([COL_PRODUCT])

    seen: set[tuple[str, ...]] = set()
    for keys in merge_chain:
        t = tuple(keys)
        if t in seen:
            continue
        seen.add(t)
        merged = _coalesce_merge(merged, keys)

    merged = merged.drop(columns=[c for c in ("_qm", "_dk") if c in merged.columns], errors="ignore")
    return pl.from_pandas(merged)


def _product_price_snapshot_from_bom_rows(rows: list[dict]) -> dict:
    """从 BOM 明细提取产品价格历史清单合并字段（供成本模拟页）；金额为 元/KG。
    优先选取报价月份最新的行；次选 CREATEDATE 最大的行；均无则取第一行。
    """
    if not rows:
        return {}

    def _row_sort_key(r: dict) -> tuple[str, str]:
        return (
            str(r.get("报价月份") or ""),
            str(r.get(COL_CREATEDATE) or ""),
        )

    r0 = max(rows, key=_row_sort_key)
    out: dict = {"价格口径": "元/KG"}
    if "产品价格" in r0:
        out["总成本"] = optional_rounded_float(r0.get("产品价格"))
    for c in _PRICE_EXTRA_COLS:
        if c in r0:
            out[c] = optional_rounded_float(r0.get(c))
    return out


@app.route("/api/upload", methods=["POST"])
def upload():
    """接收单个或多个 Excel 上传，校验列名，返回前 5 行预览数据。"""
    # 兼容多文件（files）和单文件（file）两种字段名
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        single = request.files.get("file")
        if single and single.filename:
            files = [single]
        else:
            return jsonify({"error": "未找到文件字段"}), 400

    # 校验所有文件格式（支持模板 .xls / .xlsx）
    for f in files:
        if not _is_excel_filename(f.filename or ""):
            return jsonify({
                "error": f"不支持的格式：{f.filename}",
                "hint":  "请上传 .xlsx 或 .xls 格式的 Excel（如 报价BOM历史清单）",
            }), 400

    session_id = str(uuid.uuid4())
    saved_paths: list[Path] = []
    price_path: Path | None = None

    try:
        for idx, f in enumerate(files):
            ext = _upload_save_ext(f.filename or "")
            p = UPLOAD_DIR / f"{session_id}_{idx}{ext}"
            f.save(str(p))
            saved_paths.append(p)

        price_file = request.files.get("price_file")
        has_price = False
        price_fname: str | None = None
        if price_file and price_file.filename and price_file.filename.strip():
            if not _is_excel_filename(price_file.filename):
                for p in saved_paths:
                    p.unlink(missing_ok=True)
                return jsonify({
                    "error": f"价格表格式不支持：{price_file.filename}",
                    "hint":  "产品价格表请上传 .xlsx 或 .xls（如 产品价格历史清单）",
                }), 400
            pext = _upload_save_ext(price_file.filename)
            price_path = UPLOAD_DIR / f"{session_id}_price{pext}"
            price_file.save(str(price_path))
            has_price = True
            price_fname = price_file.filename

        # 用第一个文件做列校验与预览
        # 只读 5 行数据（+表头），比读全文件快几十倍
        first_path = saved_paths[0]
        try:
            df = pd.read_excel(first_path, sheet_name=0, nrows=5)
        except Exception as exc:
            for p in saved_paths:
                p.unlink(missing_ok=True)
            if price_path is not None:
                price_path.unlink(missing_ok=True)
            return jsonify({"error": f"Excel 解析失败（第 1 个文件）：{exc}"}), 422

        df_pl   = normalize_columns(pl.from_pandas(df))
        missing = [c for c in REQUIRED_COLS if c not in df_pl.columns]
        if missing:
            for p in saved_paths:
                p.unlink(missing_ok=True)
            if price_path is not None:
                price_path.unlink(missing_ok=True)
            return jsonify({
                "error":  f"缺少必要列：{missing}",
                "detail": f"当前列名（规范化后）：{list(df_pl.columns)}",
                "hint":   (
                    "必须包含（或其 SAP 导出别名）：\n"
                    "  产品编码（或 ZMATNR）、组件编码（或 IDNRK）、"
                    "MENGE、组件单价（或 ZDJ）"
                ),
            }), 422

        # 行数：从 xlsx XML 元数据读取，几乎不耗时
        total_rows = _xlsx_row_count(first_path)

        # 保存元数据
        filenames = [f.filename for f in files]
        meta = {
            "filenames":              filenames,
            "count":                  len(files),
            "has_price":              has_price,
            "price_filename":         price_fname,
        }
        (UPLOAD_DIR / f"{session_id}_meta.json").write_text(
            json.dumps(meta, ensure_ascii=False), encoding="utf-8"
        )

        preview_rows = df.fillna("").astype(str).values.tolist()
        return jsonify({
            "session_id":              session_id,
            "filenames":               filenames,
            "file_count":              len(files),
            "has_price":               has_price,
            "price_filename":          price_fname,
            "total_rows":              total_rows,
            "columns":                 list(df.columns),
            "rows":                    preview_rows,
        })

    except Exception as exc:
        for p in saved_paths:
            p.unlink(missing_ok=True)
        if price_path and price_path.exists():
            price_path.unlink(missing_ok=True)
        for ext in (".xlsx", ".xls"):
            (UPLOAD_DIR / f"{session_id}_material_trend{ext}").unlink(missing_ok=True)
        return jsonify({"error": f"上传失败：{exc}"}), 500


@app.route("/api/session/attach_price", methods=["POST"])
def attach_session_price():
    """在已有 BOM 会话上补传第二张「产品价格表」，不重新上传 BOM。"""
    sid = (request.form.get("session_id") or "").strip()
    if not _UUID_RE.match(sid):
        return jsonify({"error": "无效的 session_id"}), 400
    meta_path = UPLOAD_DIR / f"{sid}_meta.json"
    if not meta_path.exists():
        return jsonify({
            "error": "会话不存在或已失效",
            "hint":  "请重新上传 BOM 文件",
        }), 404
    bom0 = next(UPLOAD_DIR.glob(f"{sid}_0.xlsx"), None) or next(
        UPLOAD_DIR.glob(f"{sid}_0.xls"), None
    )
    if bom0 is None or not bom0.exists():
        return jsonify({
            "error": "BOM 文件缺失",
            "hint":  "请重新上传",
        }), 404

    price_file = request.files.get("price_file")
    if not price_file or not price_file.filename or not price_file.filename.strip():
        return jsonify({"error": "未收到价格表文件"}), 400
    if not _is_excel_filename(price_file.filename):
        return jsonify({
            "error":  f"价格表格式不支持：{price_file.filename}",
            "hint":   "产品价格表请上传 .xlsx 或 .xls（如 产品价格历史清单）",
        }), 400

    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return jsonify({"error": "会话元数据损坏，请重新上传"}), 500

    price_fp = UPLOAD_DIR / f"{sid}_price{_upload_save_ext(price_file.filename)}"
    try:
        price_file.save(str(price_fp))
    except Exception as exc:
        return jsonify({"error": f"保存价格表失败：{exc}"}), 500

    meta["has_price"]      = True
    meta["price_filename"] = price_file.filename
    meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")

    return jsonify({
        "success":          True,
        "has_price":        True,
        "price_filename":   price_file.filename,
    })


@app.route("/api/session/attach_material_trend", methods=["POST"])
def attach_session_material_trend():
    """材料趋势表已停用：保留旧接口，避免旧页面请求触发 404。"""
    return jsonify({
        "success": False,
        "error": "报价材料价格趋势已停用，不再参与成本模拟分析",
    }), 410


def _sidecars_for_session(session_id: str, price_fp: Path | None) -> dict:
    ppt: dict = {}
    if price_fp is not None and price_fp.exists():
        try:
            ppt = build_product_price_timeline(price_fp)
        except Exception as exc:
            print(f"产品价格时间线构建跳过: {exc}", flush=True)
    return ppt


# ---------------------------------------------------------------------------
# API – Run analysis pipeline
# ---------------------------------------------------------------------------

# Excel 后台写入状态：token -> threading.Event（set = 写完或写失败）
_EXCEL_EVENTS: dict[str, threading.Event] = {}
_EXCEL_EVENTS_LOCK = threading.Lock()


def _bg_write_excel(
    out_path: Path,
    token: str,
    sheets: list[tuple[str, pl.DataFrame]],
) -> None:
    """后台线程：把 5 张表写入 Excel 文件，完成后 set Event 通知下载接口。
    Excel 写入是纯 IO 密集型，放后台后用户无需等待即可进入 BI 看板。
    """
    try:
        with pd.ExcelWriter(str(out_path), engine="xlsxwriter") as writer:
            for name, df in sheets:
                _write_sheet(writer, df, name)
    except Exception:
        out_path.unlink(missing_ok=True)   # 写失败则删掉残文件
    finally:
        with _EXCEL_EVENTS_LOCK:
            ev = _EXCEL_EVENTS.pop(token, None)
        if ev:
            ev.set()


def _run_pipeline(raw_df: pl.DataFrame, out_path: Path):
    """执行分析管道，返回 (lines, summary_pl, detail_pl, price_history_pl, pch, bom_pd)，Excel 写入在后台进行。

    整个流程：
      1. polars 管道计算（< 1s）
      2. 填充 BI 缓存（< 0.1s）→ 调用方完成后立即返回给前端
      3. 后台线程写 Excel（10-25s）→ 用户浏览 BI 期间在后台完成
    """
    lines           = enrich(aggregate_lines(prepare(raw_df)))
    summary_pl      = sheet_component_global_rank(lines)
    detail_pl       = sheet_bom_detail(lines)
    price_history_pl = sheet_price_history(raw_df)
    product_cost_history = build_product_cost_history(raw_df)

    # 保留规范化后的原始 BOM 数据（pandas），供新预测算法使用
    try:
        bom_pd: pd.DataFrame | None = normalize_columns(raw_df).to_pandas()
    except Exception:
        bom_pd = None

    # 提前计算所有工作表（纯内存操作，极快），存为列表传给后台线程
    sheets: list[tuple[str, pl.DataFrame]] = [
        ("产品总成本排名",     sheet_product_summary(lines)),
        ("BOM明细_占比与成本", detail_pl),
        ("组件全局成本排名",   summary_pl),
        ("跨产品用量波动",     sheet_cross_product_volatility(lines)),
        ("各产品Top3成本组件", top3_per_product(lines)),
        ("价格历史明细",       price_history_pl),
    ]

    token = out_path.stem   # 文件名即 UUID token
    if bi_cache.write_excel_enabled():
        ev = threading.Event()
        with _EXCEL_EVENTS_LOCK:
            _EXCEL_EVENTS[token] = ev
        threading.Thread(
            target=_bg_write_excel,
            args=(out_path, token, sheets),
            daemon=True,
        ).start()

    return lines, summary_pl, detail_pl, price_history_pl, product_cost_history, bom_pd


def _read_price_df(price_fp: "Path | None") -> "pd.DataFrame | None":
    """从价格文件读取 pandas DataFrame，规范化产品编码列，供新预测算法使用。"""
    if price_fp is None or not price_fp.exists():
        return None
    try:
        df = read_price_excel(price_fp)
        # 规范化产品编码列名
        if COL_PRODUCT not in df.columns:
            for alias in ("ZMATNR", "MATNR", "所属产品"):
                if alias in df.columns:
                    df = df.rename(columns={alias: COL_PRODUCT})
                    break
        # 规范化 总成本 列（兜底别名）
        if "总成本" not in df.columns:
            for alias in ("产品价格", "标价", "出厂价", "DMBTR"):
                if alias in df.columns:
                    df = df.rename(columns={alias: "总成本"})
                    break
        _, df_norm = normalize_predict_dataframes(None, df)
        return df_norm
    except Exception:
        return None


@app.route("/api/process", methods=["POST"])
def process():
    """执行 BOM 成本分析管道，将结果写为 Excel 并返回下载令牌。
    
    mode='merge'   （默认）：将所有上传文件纵向合并后生成一份报告。
    mode='per_file'         ：每个上传文件各生成一份独立报告。
    """
    data       = request.get_json(silent=True) or {}
    session_id = data.get("session_id", "")
    mode       = data.get("mode", "merge")   # "merge" | "per_file"

    if not _UUID_RE.match(session_id):
        return jsonify({"error": "无效的 session_id"}), 400

    meta_path = UPLOAD_DIR / f"{session_id}_meta.json"
    if not meta_path.exists():
        return jsonify({"error": "上传文件已过期，请重新上传"}), 404

    try:
        meta      = json.loads(meta_path.read_text(encoding="utf-8"))
        filenames = meta["filenames"]
        count     = meta["count"]
        has_price = bool(meta.get("has_price"))
        price_fp  = _upload_price_path(session_id) if has_price else None
        if has_price and price_fp is None:
            return jsonify({"error": "元数据标明有产品价格表，但文件已丢失，请重新上传"}), 404

        # ── 读取所有上传文件 ──────────────────────────────────────────────
        raw_list: list[pl.DataFrame] = []
        for i in range(count):
            p = _upload_bom_path(session_id, i)
            if p is None:
                return jsonify({"error": f"上传文件 [{filenames[i]}] 已过期，请重新上传"}), 404
            raw_list.append(_read_xlsx_as_polars(p))

        ppt = _sidecars_for_session(session_id, price_fp)

        def _cleanup_inputs() -> None:
            for i in range(count):
                for ext in (".xlsx", ".xls"):
                    (UPLOAD_DIR / f"{session_id}_{i}{ext}").unlink(missing_ok=True)
            if price_fp is not None:
                price_fp.unlink(missing_ok=True)
            for ext in (".xlsx", ".xls"):
                (UPLOAD_DIR / f"{session_id}_material_trend{ext}").unlink(missing_ok=True)
            meta_path.unlink(missing_ok=True)

        # ── 准备新预测算法所需 DataFrames ────────────────────────────────
        raw_combined_pre = pl.concat(raw_list) if len(raw_list) > 1 else raw_list[0]
        try:
            bom_pd_raw: pd.DataFrame | None = normalize_columns(raw_combined_pre).to_pandas()
        except Exception:
            bom_pd_raw = None
        price_pd_raw = _read_price_df(price_fp)
        bom_pd_raw, price_pd_raw = normalize_predict_dataframes(bom_pd_raw, price_pd_raw)
        product_categories = (
            build_product_category_index(bom_pd_raw) if bom_pd_raw is not None else {}
        )

        # ── 合并模式 ──────────────────────────────────────────────────────
        if mode == "merge":
            combined = raw_combined_pre
            if has_price and price_fp is not None:
                combined = _attach_product_prices_pl(combined, price_fp)
            out_path = RESULT_DIR / f"{session_id}.xlsx"
            lines, summary_pl, detail_pl, price_history_pl, pch, bom_pd_pipe = _run_pipeline(combined, out_path)
            bi_cache.fill_and_persist(
                session_id,
                summary_pl,
                detail_pl,
                price_history_pl,
                pch,
                ppt,
                bom_pd=bom_pd_raw,
                price_pd=price_pd_raw,
                product_categories=product_categories,
            )

            _cleanup_inputs()

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            return jsonify({
                "success":        True,
                "mode":           "merge",
                "download_token": session_id,
                "filename":       f"bom_analysis_{ts}.xlsx",
                "stats": {
                    "products":   lines[COL_PRODUCT].n_unique(),
                    "components": lines[COL_COMPONENT].n_unique(),
                    "total_rows": lines.height,
                },
            })

        # ── 分别分析模式 ──────────────────────────────────────────────────
        else:
            tokens: list[dict] = []
            total_rows = 0

            for i, (raw_df_item, fname) in enumerate(zip(raw_list, filenames)):
                token    = str(uuid.uuid4())
                out_path = RESULT_DIR / f"{token}.xlsx"
                try:
                    bom_pd_item: pd.DataFrame | None = normalize_columns(raw_df_item).to_pandas()
                except Exception:
                    bom_pd_item = None
                bom_pd_item, _ = normalize_predict_dataframes(bom_pd_item, None)
                product_categories_item = (
                    build_product_category_index(bom_pd_item) if bom_pd_item is not None else {}
                )
                if has_price and price_fp is not None:
                    raw_df_item = _attach_product_prices_pl(raw_df_item, price_fp)
                lines, summary_pl, detail_pl, price_history_pl, pch, _ = _run_pipeline(raw_df_item, out_path)
                bi_cache.fill_and_persist(
                    token,
                    summary_pl,
                    detail_pl,
                    price_history_pl,
                    pch,
                    ppt,
                    bom_pd=bom_pd_item,
                    price_pd=price_pd_raw,
                    product_categories=product_categories_item,
                )
                total_rows += lines.height
                tokens.append({
                    "token":    token,
                    "filename": fname,
                    "stats": {
                        "products":   lines[COL_PRODUCT].n_unique(),
                        "components": lines[COL_COMPONENT].n_unique(),
                        "total_rows": lines.height,
                    },
                })

            _cleanup_inputs()

            return jsonify({
                "success":        True,
                "mode":           "per_file",
                "download_tokens": tokens,
                "stats": {
                    "files":      len(tokens),
                    "total_rows": total_rows,
                },
            })

    except Exception as exc:
        tb_lines = traceback.format_exc().strip().split("\n")
        detail   = tb_lines[-1] if tb_lines else None
        return jsonify({"error": f"分析失败：{exc}", "detail": detail}), 500



def _parse_excel_dates(series: pd.Series) -> pd.Series:
    """把 Excel/SAP 常见日期格式统一成 pandas datetime。"""
    text = (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str[:8]
    )
    return pd.to_datetime(text, format="%Y%m%d", errors="coerce")


def _first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    return None


class AiPurchaseAdviceUnavailable(Exception):
    """未配置 DeepSeek 等前置条件，AI 采购建议不可用。"""


def _deepseek_api_key() -> str:
    return os.environ.get("DEEPSEEK_API_KEY", "").strip()


def _baidu_search_api_key() -> str:
    return os.environ.get("BAIDU_SEARCH_API_KEY", "").strip()


def _external_capabilities() -> dict:
    """外网 / API Key 能力（与会话无关）。"""
    deepseek = bool(_deepseek_api_key())
    baidu = bool(_baidu_search_api_key())
    ai_reason = ""
    if not deepseek:
        ai_reason = (
            "未配置 DEEPSEEK_API_KEY。请在 web 目录下创建 .env 并设置该变量后重启服务。"
        )
    baidu_note = ""
    if deepseek and not baidu:
        baidu_note = (
            "未配置 BAIDU_SEARCH_API_KEY，联网行情搜索不可用；"
            "仍将基于本地 Excel 价格波动与模型常识生成建议（证据可能较弱）。"
        )
    return {
        "ai_purchase_advice": {
            "enabled": deepseek,
            "reason": ai_reason,
            "baidu_search": baidu,
            "baidu_note": baidu_note,
        },
    }


init_bi_blueprint(_external_capabilities)
app.register_blueprint(bi_bp)


def _require_deepseek_for_ai() -> None:
    if not _deepseek_api_key():
        raise AiPurchaseAdviceUnavailable(
            "未配置 DEEPSEEK_API_KEY。请在 web/.env 中设置后重启服务。"
        )


def _deepseek_model() -> str:
    return os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-pro").strip() or "deepseek-v4-pro"


def _deepseek_headers() -> dict:
    api_key = _deepseek_api_key()
    if not api_key:
        raise AiPurchaseAdviceUnavailable(
            "未配置 DEEPSEEK_API_KEY。请在 web/.env 中设置后重启服务。"
        )
    return {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }


def _requests_post_json(
    url: str,
    payload: dict,
    headers: dict,
    *,
    timeout: tuple[float, float] | float,
) -> dict:
    try:
        resp = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=timeout,
        )
        resp.raise_for_status()
        return resp.json()
    except requests.HTTPError as exc:
        detail = (exc.response.text if exc.response is not None else str(exc))[:300]
        code = exc.response.status_code if exc.response is not None else "?"
        raise RuntimeError(f"API 调用失败（HTTP {code}）：{detail}") from exc
    except requests.RequestException as exc:
        raise RuntimeError(f"无法连接 API：{exc}") from exc


def _requests_stream_sse_deltas(
    url: str,
    payload: dict,
    headers: dict,
    *,
    connect_timeout: float = 10.0,
    read_timeout: float = 180.0,
):
    """解析 OpenAI 风格 SSE，逐块 yield delta content。"""
    try:
        with requests.post(
            url,
            json=payload,
            headers=headers,
            stream=True,
            timeout=(connect_timeout, read_timeout),
        ) as resp:
            resp.raise_for_status()
            for raw_line in resp.iter_lines(decode_unicode=True):
                if not raw_line:
                    continue
                line = raw_line.strip()
                if not line.startswith("data:"):
                    continue
                data = line[5:].strip()
                if data == "[DONE]":
                    break
                try:
                    chunk = json.loads(data)
                    delta = chunk["choices"][0].get("delta", {}).get("content", "")
                except Exception:
                    delta = ""
                if delta:
                    yield delta
    except requests.HTTPError as exc:
        detail = (exc.response.text if exc.response is not None else str(exc))[:300]
        code = exc.response.status_code if exc.response is not None else "?"
        raise RuntimeError(f"DeepSeek API 调用失败（HTTP {code}）：{detail}") from exc
    except requests.RequestException as exc:
        raise RuntimeError(f"无法连接 DeepSeek API：{exc}") from exc


def _deepseek_chat_completion(payload: dict, *, timeout: int = 60) -> dict:
    read_t = float(timeout) if timeout else 90.0
    return _requests_post_json(
        "https://api.deepseek.com/chat/completions",
        payload,
        _deepseek_headers(),
        timeout=(10.0, read_t),
    )


_MARKET_SEARCH_TOOL = {
    "type": "function",
    "function": {
        "name": "search_market_price",
        "description": "搜索某个物料或上游原材料最近三个月的市场价格、行情、供需、检修、成本变化等公开网页信息。",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "搜索关键词，例如：PPR 管件 最近三个月 PP 原料 价格 走势 检修 开工率",
                },
                "material_name": {
                    "type": "string",
                    "description": "该搜索对应的物料名称或上游原料名称。",
                },
            },
            "required": ["query", "material_name"],
        },
    },
}


def _run_market_search_tool(query: str, material_name: str) -> dict:
    results = web_search_material_price(query, max_results=4)
    evidence_strength = "较弱"
    if any(item.get("snippet") for item in results):
        evidence_strength = "中等"
    if any(re.search(r"价格|行情|报价|走势|涨|跌|检修|开工|成本|原油|PP|PE|PPR|钛白粉|碳酸钙", item.get("title", "") + item.get("snippet", ""), re.I) for item in results):
        evidence_strength = "可用于方向判断"
    return {
        "query": query,
        "material_name": material_name,
        "evidence_strength": evidence_strength,
        "results": results,
    }


def _prepare_market_search_messages(
    materials: list[dict],
    start_date: str,
    end_date: str,
) -> tuple[list[dict], list[dict]]:
    """让 DeepSeek 先决定搜索什么，再执行工具并返回可用于最终回答的消息。"""
    first_payload = _deepseek_purchase_advice_payload(
        materials,
        start_date,
        end_date,
        stream=False,
        use_tools=True,
    )
    messages = list(first_payload["messages"])
    first_resp = _deepseek_chat_completion(first_payload, timeout=90)
    message = first_resp["choices"][0]["message"]
    tool_calls = message.get("tool_calls") or []

    if not tool_calls:
        messages.append({
            "role": "assistant",
            "content": message.get("content") or "未生成搜索工具调用。",
        })
        messages.append({
            "role": "user",
            "content": "你没有调用搜索工具。请明确说明实时行情证据不足，并基于产业链通用逻辑给出保守采购建议。",
        })
        return messages, []

    messages.append({
        "role": "assistant",
        "content": message.get("content"),
        "tool_calls": tool_calls,
    })

    evidence: list[dict] = []
    for tool_call in tool_calls[:8]:
        fn = tool_call.get("function", {})
        if fn.get("name") != "search_market_price":
            continue
        try:
            args = json.loads(fn.get("arguments") or "{}")
        except json.JSONDecodeError:
            args = {}
        query = str(args.get("query") or "").strip()
        material_name = str(args.get("material_name") or query or "未知物料").strip()
        if not query:
            query = f"{material_name} 最近三个月 价格 行情 走势"

        result = _run_market_search_tool(query, material_name)
        evidence.append(result)
        messages.append({
            "role": "tool",
            "tool_call_id": tool_call["id"],
            "content": json.dumps(result, ensure_ascii=False),
        })

    messages.append({
        "role": "user",
        "content": (
            "请基于以上工具搜索结果输出最终 Markdown 采购行情分析。"
            "必须先总结市场因素和短期展望，再给推荐购入、涨幅需注意和采购操作建议。"
            "如果搜索结果包含价格、报价、行情、走势、上游原料等关键词，就可以作为方向性判断依据。"
            "不要反复写“实时行情证据不足”；只在完全没有相关结果的物料上标注证据不足。"
            "如果没有精确涨跌幅，就写“未获得精确涨跌幅，按方向性行情判断”。"
        ),
    })
    return messages, evidence


def _call_deepseek_for_purchase_advice(materials: list[dict], start_date: str, end_date: str) -> str:
    messages, _ = _prepare_market_search_messages(materials, start_date, end_date)
    payload = _deepseek_purchase_advice_payload(
        materials,
        start_date,
        end_date,
        stream=False,
        use_tools=False,
        thinking_enabled=True,
        messages=messages,
    )
    data = _deepseek_chat_completion(payload, timeout=90)
    try:
        return data["choices"][0]["message"]["content"].strip()
    except Exception as exc:
        raise RuntimeError(f"DeepSeek 返回格式异常：{data}") from exc


def _build_purchase_advice_materials(session_id: str, file_path: Path) -> dict:
    cache = bi_cache.get(session_id, file_path)
    price_history = cache.get("price_history", [])
    if not price_history:
        raise ValueError("无物料明细数据，请重新上传并执行分析")

    df = pd.DataFrame(price_history)
    name_col = "组件名称" if "组件名称" in df.columns else ("MAKTX" if "MAKTX" in df.columns else None)
    required = [COL_COMPONENT]
    missing = [c for c in required if c not in df.columns]
    if missing or not name_col:
        raise ValueError(
            "缺少 AI 联网分析所需列：需要 组件编码、MAKTX（或组件名称）；"
            f"当前缺少：{missing + ([] if name_col else ['MAKTX'])}"
        )

    df[name_col] = df[name_col].fillna("").astype(str).str.strip()
    df[COL_COMPONENT] = df[COL_COMPONENT].fillna("").astype(str).str.strip()
    df = df[(df[name_col] != "") & (df[COL_COMPONENT] != "")]
    if df.empty:
        raise ValueError("MAKTX / 组件编码有效数据为空，无法生成 AI 建议")

    rows: list[dict] = []
    for (comp, name), grp in df.groupby([COL_COMPONENT, name_col], sort=False):
        row = {
            "组件编码": str(comp),
            "物料名称": str(name),
            "记录数": int(len(grp)),
            "涉及产品数": int(grp[COL_PRODUCT].nunique()) if COL_PRODUCT in grp.columns else 0,
        }
        rows.append(row)

    if not rows:
        raise ValueError("未找到可分析的 MAKTX 物料记录")

    # 先选使用面更广、记录更多的物料交给 DeepSeek，后端不再做联网搜索。
    rows.sort(key=lambda r: (r["涉及产品数"], r["记录数"]), reverse=True)
    selected = rows[:12]

    today = datetime.now()
    start_dt = today - pd.DateOffset(months=3)
    return {
        "materials": selected,
        "date_range": {"start": start_dt.strftime("%Y-%m-%d"), "end": today.strftime("%Y-%m-%d")},
        "total": len(rows),
    }


def _deepseek_purchase_advice_payload(
    materials: list[dict],
    start_date: str,
    end_date: str,
    *,
    stream: bool,
    use_tools: bool = False,
    thinking_enabled: bool = False,
    messages: list[dict] | None = None,
) -> dict:
    prompt = (
        "你是制造业采购和原材料成本分析顾问。Excel 只提供待分析物料清单，不提供价格结论。"
        "请你基于最近三个月的公开市场行情、网上价格趋势、原材料供需信息，"
        "判断这些物料是否适合购入或需要注意涨价风险。"
        "输出风格要像专业采购行情简报：先讲市场逻辑，再给动作建议，不能只罗列物料。\n"
        "要求：\n"
        "1. 用 Markdown 输出。\n"
        "2. 必须按以下结构输出：\n"
        "   - `## 近期影响市场价格的核心因素`\n"
        "   - `## 短期市场展望`\n"
        "   - `## 推荐购入`\n"
        "   - `## 涨幅较大需注意`\n"
        "   - `## 采购操作建议`\n"
        "3. `近期影响市场价格的核心因素` 至少写 3 条，围绕供应、成本、需求、替代品、进口/出口、季节性、政策或检修等因素展开。\n"
        "4. `短期市场展望` 要给出明确判断，例如高位震荡、偏强运行、弱势回调、低位企稳，并说明为什么。\n"
        "5. `推荐购入` 和 `涨幅较大需注意` 每条包含物料名称、组件编码、判断方向、依据、建议动作。\n"
        "6. 每条建议必须有“有理有据”的原因：例如上游原料、供需变化、成本支撑、需求淡旺季、库存或行业开工率等。\n"
        "7. 如果工具搜索结果没有精确涨跌幅，不要反复写“实时行情证据不足”；应改写为“未获得精确涨跌幅，以下为方向性判断”。\n"
        "8. 只在某个物料完全没有价格/行情/报价/走势相关搜索结果时，才对该物料标注“证据不足”。\n"
        "9. 可以把搜索结果中的价格页、报价页、行情页、原料价格走势页作为方向性证据，但不要编造具体涨跌幅、具体开工率或具体价格。\n"
        "10. 不要按 Excel 记录数排序，也不要假装 Excel 里有市场涨跌幅。\n"
        "11. 优先结合物料关键词判断，例如 PE、PP-R、管材、管件、包装膜、钛白粉、碳酸钙等相关大宗原料趋势。\n"
        "12. 语言要像采购经理能直接看的行情简报：简短、有判断、有依据、有动作。\n\n"
        "工具使用要求：\n"
        "- 在最终分析前，优先调用 `search_market_price` 工具搜索 3-8 个关键行情问题。\n"
        "- 搜索词由你根据物料清单自行设计，既要搜具体物料，也要搜上游原料，如 PP、PE、PPR、钛白粉、碳酸钙等。\n"
        "- 最终结论要引用工具结果里的标题/摘要关键词作为依据。\n\n"
        "参考表达风格：\n"
        "- 供应收缩：若上游装置检修、开工率下降或供应偏紧，价格下方支撑增强。\n"
        "- 高成本支撑：若原油、丙烯、丙烷、甲醇、钛矿等上游成本处于高位，下游物料存在被动跟涨压力。\n"
        "- 需求端压力：若处于传统淡季或终端开工不足，价格上行空间会受压。\n"
        "- 操作建议：不要只写“推荐购买”，要写按需补库、锁价、分批采购、观望、供应商比价或替代料评估。\n\n"
        f"日期范围：{start_date} 至 {end_date}\n"
        f"待分析物料清单 JSON：{json.dumps(materials, ensure_ascii=False)}"
    )
    payload = {
        "model": _deepseek_model(),
        "thinking": {"type": "enabled" if thinking_enabled else "disabled"},
        "temperature": 0.2,
        "stream": stream,
        "messages": messages or [
            {"role": "system", "content": "你只基于用户提供的数据给采购建议，避免空泛表达。"},
            {"role": "user", "content": prompt},
        ],
    }
    if use_tools:
        payload["tools"] = [_MARKET_SEARCH_TOOL]
        payload["tool_choice"] = "auto"
    return payload


def _stream_deepseek_for_purchase_advice(materials: list[dict], start_date: str, end_date: str):
    messages, _ = _prepare_market_search_messages(materials, start_date, end_date)

    payload = _deepseek_purchase_advice_payload(
        materials,
        start_date,
        end_date,
        stream=True,
        use_tools=False,
        thinking_enabled=True,
        messages=messages,
    )
    yield from _requests_stream_sse_deltas(
        "https://api.deepseek.com/chat/completions",
        payload,
        _deepseek_headers(),
        connect_timeout=10.0,
        read_timeout=180.0,
    )


@app.route("/api/bi/ai_purchase_advice/<session_id>")
def bi_ai_purchase_advice(session_id: str):
    """读取 MAKTX 物料名称，计算最近三个月价格波动，并调用 DeepSeek 生成采购建议。"""
    if not _UUID_RE.match(session_id):
        return jsonify({"error": "无效的令牌"}), 400

    try:
        _require_deepseek_for_ai()
    except AiPurchaseAdviceUnavailable as exc:
        return jsonify({
            "error": str(exc),
            "code": "AI_DISABLED",
            "enabled": False,
        }), 503

    if not bi_cache.session_ready(session_id):
        return jsonify({"error": "会话不存在或已过期，请重新分析"}), 404

    file_path = RESULT_DIR / f"{session_id}.xlsx"
    try:
        info = _build_purchase_advice_materials(session_id, file_path)
        materials_for_ai = info["materials"]
        start_s = info["date_range"]["start"]
        end_s = info["date_range"]["end"]
        advice = _call_deepseek_for_purchase_advice(materials_for_ai, start_s, end_s)

        return jsonify({
            "advice": advice,
            "materials": materials_for_ai,
            "date_range": info["date_range"],
            "total": info["total"],
        })

    except AiPurchaseAdviceUnavailable as exc:
        return jsonify({"error": str(exc), "code": "AI_DISABLED", "enabled": False}), 503
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 422
    except Exception as exc:
        tb = traceback.format_exc().strip().split("\n")
        return jsonify({"error": f"AI 采购建议生成失败：{exc}", "detail": tb[-1]}), 500


@app.route("/api/bi/ai_purchase_advice_stream/<session_id>")
def bi_ai_purchase_advice_stream(session_id: str):
    """流式返回 Markdown 格式 AI 采购建议。"""
    if not _UUID_RE.match(session_id):
        return jsonify({"error": "无效的令牌"}), 400

    try:
        _require_deepseek_for_ai()
    except AiPurchaseAdviceUnavailable as exc:
        return jsonify({
            "error": str(exc),
            "code": "AI_DISABLED",
            "enabled": False,
        }), 503

    file_path = RESULT_DIR / f"{session_id}.xlsx"
    if not bi_cache.session_ready(session_id):
        return jsonify({"error": "会话不存在或已过期，请重新分析"}), 404

    def sse(event: str, data) -> str:
        return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

    @stream_with_context
    def generate():
        try:
            info = _build_purchase_advice_materials(session_id, file_path)
            yield sse("meta", info)
            for delta in _stream_deepseek_for_purchase_advice(
                info["materials"],
                info["date_range"]["start"],
                info["date_range"]["end"],
            ):
                yield sse("delta", {"text": delta})
            yield sse("done", {"ok": True})
        except AiPurchaseAdviceUnavailable as exc:
            yield sse("error", {
                "error": str(exc),
                "code": "AI_DISABLED",
                "enabled": False,
            })
        except Exception as exc:
            yield sse("error", {"error": f"AI 采购建议生成失败：{exc}"})

    return Response(generate(), mimetype="text/event-stream")


# ---------------------------------------------------------------------------
# API – Download result
# ---------------------------------------------------------------------------

@app.route("/api/download/<session_id>")
def download(session_id: str):
    """凭 session_id（下载令牌）下载分析结果 Excel。
    若 Excel 仍在后台写入，最多等待 120 秒；写完即返回文件。
    """
    if not _UUID_RE.match(session_id):
        return jsonify({"error": "无效的下载令牌"}), 400

    if not bi_cache.write_excel_enabled():
        return jsonify({
            "error": "Excel 导出已关闭（BOM_WRITE_EXCEL=0）。请使用 BI 看板查看分析结果。",
        }), 503

    file_path = RESULT_DIR / f"{session_id}.xlsx"

    # 如果后台写入尚未完成，等 Event 信号（最多 120s）
    with _EXCEL_EVENTS_LOCK:
        ev = _EXCEL_EVENTS.get(session_id)
    if ev is not None:
        ev.wait(timeout=120)

    if not file_path.exists():
        return jsonify({"error": "文件生成失败或已过期，请重新分析"}), 404

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"bom_analysis_{ts}.xlsx",
    )


# ---------------------------------------------------------------------------
# API – Excel ready status（供前端判断导出按钮是否可用）
# ---------------------------------------------------------------------------

@app.route("/api/excel_ready/<session_id>")
def excel_ready(session_id: str):
    """返回 Excel 是否已写好。前端可轮询此接口来更新「导出」按钮状态。"""
    if not _UUID_RE.match(session_id):
        return jsonify({"ready": False}), 400
    with _EXCEL_EVENTS_LOCK:
        ev = _EXCEL_EVENTS.get(session_id)
    if ev is None:
        # 不在写入队列里：要么已写完（文件存在），要么根本没有
        ready = (RESULT_DIR / f"{session_id}.xlsx").exists()
    else:
        ready = ev.is_set()
    return jsonify({"ready": ready})


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def _open_browser_when_ready(url: str = "http://127.0.0.1:5000/", delay_sec: float = 1.0) -> None:
    """服务启动后自动打开默认浏览器。设环境变量 BOM_WEB_NO_BROWSER=1 可关闭。"""
    flag = os.environ.get("BOM_WEB_NO_BROWSER", "").strip().lower()
    if flag in ("1", "true", "yes", "on"):
        return
    time.sleep(delay_sec)
    webbrowser.open(url)


def _should_auto_open_browser(*, debug: bool, use_reloader: bool) -> bool:
    """是否在「当前进程」里调度打开浏览器（避免 debug 重载时父、子进程各开一次）。"""
    if os.environ.get("BOM_WEB_NO_BROWSER", "").strip().lower() in ("1", "true", "yes", "on"):
        return False
    # 在 app.run() 之前 app.debug 仍为默认 False，不能用 app.debug 判断。
    if debug and use_reloader:
        return os.environ.get("WERKZEUG_RUN_MAIN") == "true"
    return True


def _should_run_startup_tasks(*, debug: bool, use_reloader: bool) -> bool:
    if debug and use_reloader:
        return os.environ.get("WERKZEUG_RUN_MAIN") == "true"
    return True


if __name__ == "__main__":
    _DEBUG = os.environ.get("FLASK_DEBUG", "").strip().lower() in ("1", "true")
    _USE_RELOADER = _DEBUG

    if _should_run_startup_tasks(debug=_DEBUG, use_reloader=_USE_RELOADER):
        _purge_upload_and_results()
        _schedule_daily_purge()

    print()
    print("=" * 52)
    print("   BOM 成本分析 Web 应用")
    print("   访问地址 → http://127.0.0.1:5000")
    print("=" * 52)
    print()
    if _should_auto_open_browser(debug=_DEBUG, use_reloader=_USE_RELOADER):
        threading.Thread(
            target=_open_browser_when_ready,
            kwargs={"url": "http://127.0.0.1:5000/", "delay_sec": 1.0},
            daemon=True,
        ).start()
    _HOST = os.environ.get("BOM_WEB_HOST", "127.0.0.1")
    app.run(debug=_DEBUG, use_reloader=_USE_RELOADER, port=5000, host=_HOST)
